from fastapi import FastAPI, Request, Query, Form
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from datetime import datetime, timedelta
from collections import defaultdict
from io import BytesIO
from sheets.client import sheets
from config.settings import settings
from zoneinfo import ZoneInfo
import secrets
import uvicorn

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    Workbook = None

app = FastAPI(title="Merchant Care Analytics")
templates = Jinja2Templates(directory="dashboard/templates")

TZ = ZoneInfo("Asia/Tashkent")

# Простые сессии в памяти: token → "view" | "admin"
_sessions: dict[str, str] = {}


def _auth_enabled() -> bool:
    return bool(settings.dashboard_view_password)


def _get_role(request: Request) -> str | None:
    """None = не авторизован, view / admin"""
    if not _auth_enabled():
        return "admin"  # без пароля — полный доступ (как раньше)

    token = request.cookies.get("mca_session")
    if not token:
        return None
    return _sessions.get(token)


def _require_login(request: Request) -> RedirectResponse | None:
    if _get_role(request) is None:
        return RedirectResponse(url="/login", status_code=303)
    return None


def parse_date(s: str):
    try:
        return datetime.strptime(s.strip(), "%d.%m.%Y").date()
    except Exception:
        return None


def date_in_range(date_str: str, start, end) -> bool:
    d = parse_date(date_str)
    if not d:
        return False
    return start <= d <= end


def get_period_bounds(period: str, date_from: str | None = None, date_to: str | None = None):
    today = datetime.now(TZ).date()

    if period == "custom" and date_from and date_to:
        start = parse_date(date_from)
        end = parse_date(date_to)
        if start and end and start <= end:
            return start, end, f"{date_from} — {date_to}"

    if period == "month":
        start = today.replace(day=1)
        end = today
        months = {
            1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
            5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
            9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
        }
        label = f"{months.get(start.month, '')} {start.year}"
        return start, end, label

    start = today - timedelta(days=today.weekday())
    end = today
    label = f"{start.strftime('%d.%m.%Y')} — {end.strftime('%d.%m.%Y')}"
    return start, end, label


def build_report(start, end, label: str):
    sheets.connect()
    cases = sheets.get_all_cases()
    messages = sheets.get_all_messages()

    total_cases = 0
    closed_cases = 0
    open_cases = 0
    response_times = []
    sla_ok = 0
    sla_total = 0

    employee_stats = defaultdict(lambda: {
        "count": 0,
        "response_times": [],
        "sla_ok": 0,
        "sla_total": 0,
    })
    merchant_cases = defaultdict(int)
    merchant_messages = defaultdict(int)
    messages_total = 0

    for case in cases:
        case_date = str(case.get("Дата", "")).strip()
        if not date_in_range(case_date, start, end):
            continue

        status = str(case.get("Status", "")).strip()
        employee = str(case.get("Employee", "")).strip()
        response_time = case.get("Response Time", "")
        sla = str(case.get("SLA", "")).strip()
        merchant = str(case.get("Merchant", "")).strip()

        total_cases += 1
        if status == "Closed":
            closed_cases += 1
        if status == "Open":
            open_cases += 1
        if merchant:
            merchant_cases[merchant] += 1

        if response_time not in ("", None):
            try:
                rt = int(float(response_time))
                response_times.append(rt)
                sla_total += 1
                if sla == "OK":
                    sla_ok += 1
            except Exception:
                pass

        if employee:
            employee_stats[employee]["count"] += 1
            if response_time not in ("", None):
                try:
                    rt = int(float(response_time))
                    employee_stats[employee]["response_times"].append(rt)
                    employee_stats[employee]["sla_total"] += 1
                    if sla == "OK":
                        employee_stats[employee]["sla_ok"] += 1
                except Exception:
                    pass

    for msg in messages:
        msg_date = str(msg.get("Дата", "")).strip()
        if not date_in_range(msg_date, start, end):
            continue
        messages_total += 1
        merchant = str(msg.get("Merchant", "")).strip()
        if merchant:
            merchant_messages[merchant] += 1

    avg_response = round(sum(response_times) / len(response_times), 1) if response_times else 0
    sla_percent = round(sla_ok / sla_total * 100, 1) if sla_total else 0

    emp_kpi = []
    for name, stats in employee_stats.items():
        avg = round(
            sum(stats["response_times"]) / len(stats["response_times"]), 1
        ) if stats["response_times"] else 0
        sla_p = round(
            stats["sla_ok"] / stats["sla_total"] * 100, 1
        ) if stats["sla_total"] else 0
        emp_kpi.append({
            "name": name,
            "count": stats["count"],
            "avg_response": avg,
            "sla": sla_p,
        })
    emp_kpi.sort(key=lambda x: x["count"], reverse=True)

    best_employee = emp_kpi[0]["name"] if emp_kpi else "—"
    with_sla = [e for e in emp_kpi if e["sla"] > 0]
    if with_sla:
        best_employee = max(with_sla, key=lambda x: (x["sla"], -x["avg_response"]))["name"]

    top_merchants = sorted(
        [{"name": k, "messages": merchant_messages.get(k, 0), "cases": merchant_cases.get(k, 0)}
         for k in set(list(merchant_messages.keys()) + list(merchant_cases.keys()))],
        key=lambda x: x["messages"] or x["cases"],
        reverse=True,
    )[:15]

    top_load = top_merchants[0]["name"] if top_merchants else "—"

    return {
        "label": label,
        "date_from": start.strftime("%d.%m.%Y"),
        "date_to": end.strftime("%d.%m.%Y"),
        "messages_total": messages_total,
        "total_cases": total_cases,
        "closed_cases": closed_cases,
        "open_cases": open_cases,
        "avg_response": avg_response,
        "sla_percent": sla_percent,
        "best_employee": best_employee,
        "top_load": top_load,
        "emp_kpi": emp_kpi,
        "top_merchants": top_merchants,
        "now": datetime.now(TZ).strftime("%d.%m.%Y %H:%M"),
    }


def get_dashboard_data(filter_date: str | None = None):
    sheets.connect()
    cases = sheets.get_all_cases()
    messages = sheets.get_all_messages()

    today = datetime.now(TZ).strftime("%d.%m.%Y")
    selected = filter_date if filter_date else today

    total_today = 0
    closed_today = 0
    open_count = 0
    response_times = []
    sla_ok = 0
    sla_total = 0

    employee_stats = defaultdict(lambda: {
        "count": 0,
        "response_times": [],
        "sla_ok": 0,
        "sla_total": 0,
    })

    recent_cases = []
    open_cases = []

    for case in cases:
        status = str(case.get("Status", "")).strip()
        case_date = str(case.get("Дата", "")).strip()
        employee = str(case.get("Employee", "")).strip()
        response_time = case.get("Response Time", "")
        sla = str(case.get("SLA", "")).strip()

        if case_date == selected:
            total_today += 1
            if status == "Closed":
                closed_today += 1

        if status == "Open":
            open_count += 1
            open_cases.append(case)

        if case_date == selected and response_time not in ("", None):
            try:
                rt = int(float(response_time))
                response_times.append(rt)
                sla_total += 1
                if sla == "OK":
                    sla_ok += 1
            except Exception:
                pass

        if case_date == selected and employee:
            employee_stats[employee]["count"] += 1
            if response_time not in ("", None):
                try:
                    rt = int(float(response_time))
                    employee_stats[employee]["response_times"].append(rt)
                    employee_stats[employee]["sla_total"] += 1
                    if sla == "OK":
                        employee_stats[employee]["sla_ok"] += 1
                except Exception:
                    pass

        recent_cases.append(case)

    messages_today = 0
    merchant_messages = defaultdict(int)
    hours = defaultdict(int)

    for msg in messages:
        msg_date = str(msg.get("Дата", "")).strip()
        msg_time = str(msg.get("Время", "")).strip()
        merchant = str(msg.get("Merchant", "")).strip()

        if msg_date != selected:
            continue

        messages_today += 1
        if merchant:
            merchant_messages[merchant] += 1
        try:
            hour = int(msg_time.split(":")[0])
            hours[hour] += 1
        except Exception:
            pass

    avg_response = round(sum(response_times) / len(response_times), 1) if response_times else 0
    sla_percent = round(sla_ok / sla_total * 100, 1) if sla_total else 0

    emp_kpi = []
    for name, stats in employee_stats.items():
        avg = round(sum(stats["response_times"]) / len(stats["response_times"]), 1) if stats["response_times"] else 0
        sla_p = round(stats["sla_ok"] / stats["sla_total"] * 100, 1) if stats["sla_total"] else 0
        emp_kpi.append({
            "name": name,
            "count": stats["count"],
            "avg_response": avg,
            "sla": sla_p,
        })
    emp_kpi.sort(key=lambda x: x["count"], reverse=True)

    top_merchants = sorted(
        [{"name": k, "count": v} for k, v in merchant_messages.items()],
        key=lambda x: x["count"],
        reverse=True,
    )[:10]

    hourly = []
    max_h = max(hours.values()) if hours else 1
    for h in range(8, 21):
        cnt = hours.get(h, 0)
        bar = int(cnt / max_h * 20) if max_h else 0
        hourly.append({"hour": f"{h:02d}", "count": cnt, "bar": "█" * bar + "░" * (20 - bar)})

    recent_cases = list(reversed(recent_cases))[:15]
    open_cases = list(reversed(open_cases))[:10]
    lunch_start, lunch_end = sheets.get_lunch()

    return {
        "total_today": total_today,
        "closed_today": closed_today,
        "open_count": open_count,
        "avg_response": avg_response,
        "sla_percent": sla_percent,
        "messages_today": messages_today,
        "emp_kpi": emp_kpi,
        "top_merchants": top_merchants,
        "hourly": hourly,
        "recent_cases": recent_cases,
        "open_cases": open_cases,
        "now": datetime.now(TZ).strftime("%d.%m.%Y %H:%M"),
        "selected_date": selected,
        "lunch_start": lunch_start,
        "lunch_end": lunch_end,
    }


@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request, error: str | None = None):
    if not _auth_enabled():
        return RedirectResponse(url="/", status_code=303)
    if _get_role(request):
        return RedirectResponse(url="/", status_code=303)
    return templates.TemplateResponse(
        "login.html",
        {
            "request": request,
            "error": error,
            "title": settings.dashboard_title,
        },
    )


@app.post("/login")
async def login_submit(request: Request, password: str = Form(...)):
    view_pw = settings.dashboard_view_password or ""
    admin_pw = settings.dashboard_admin_password or view_pw

    role = None
    if password == admin_pw and admin_pw:
        role = "admin"
    elif password == view_pw and view_pw:
        role = "view"

    if not role:
        return templates.TemplateResponse(
            "login.html",
            {
                "request": request,
                "error": "Неверный пароль",
                "title": settings.dashboard_title,
            },
            status_code=401,
        )

    token = secrets.token_hex(16)
    _sessions[token] = role

    resp = RedirectResponse(url="/", status_code=303)
    resp.set_cookie(
        key="mca_session",
        value=token,
        httponly=True,
        samesite="lax",
        max_age=60 * 60 * 12,
    )
    return resp


@app.get("/logout")
async def logout(request: Request):
    token = request.cookies.get("mca_session")
    if token and token in _sessions:
        del _sessions[token]
    resp = RedirectResponse(url="/login", status_code=303)
    resp.delete_cookie("mca_session")
    return resp


@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request, date: str | None = Query(None)):
    redirect = _require_login(request)
    if redirect:
        return redirect

    role = _get_role(request)
    data = get_dashboard_data(date)
    data["is_admin"] = role == "admin"
    data["auth_enabled"] = _auth_enabled()
    data["title"] = settings.dashboard_title
    return templates.TemplateResponse("index.html", {"request": request, **data})


@app.get("/reports", response_class=HTMLResponse)
async def reports(
    request: Request,
    period: str = Query("week"),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
):
    redirect = _require_login(request)
    if redirect:
        return redirect

    start, end, label = get_period_bounds(period, date_from, date_to)
    data = build_report(start, end, label)
    data.update({
        "period": period,
        "date_from_input": date_from or start.strftime("%d.%m.%Y"),
        "date_to_input": date_to or end.strftime("%d.%m.%Y"),
        "auth_enabled": _auth_enabled(),
        "title": settings.dashboard_title,
    })
    return templates.TemplateResponse("reports.html", {"request": request, **data})


@app.get("/reports/excel")
async def reports_excel(
    request: Request,
    period: str = Query("week"),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
):
    redirect = _require_login(request)
    if redirect:
        return redirect

    if Workbook is None:
        return HTMLResponse("Установите openpyxl: pip install openpyxl", status_code=500)

    start, end, label = get_period_bounds(period, date_from, date_to)
    data = build_report(start, end, label)

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    header_font = Font(bold=True, size=14)
    bold = Font(bold=True)

    ws["A1"] = f"{settings.dashboard_title} — отчёт"
    ws["A1"].font = header_font
    ws["A2"] = f"Период: {label}"
    ws["A3"] = f"Сформирован: {data['now']}"

    ws["A5"] = "Показатель"
    ws["B5"] = "Значение"
    ws["A5"].font = bold
    ws["B5"].font = bold

    rows = [
        ("Сообщений (нагрузка)", data["messages_total"]),
        ("Cases всего", data["total_cases"]),
        ("Закрыто", data["closed_cases"]),
        ("Открыто", data["open_cases"]),
        ("Средний ответ (мин)", data["avg_response"]),
        ("SLA %", data["sla_percent"]),
        ("Лучший сотрудник", data["best_employee"]),
        ("Самый загруженный партнёр", data["top_load"]),
    ]
    for i, (k, v) in enumerate(rows, start=6):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v

    ws["A15"] = "KPI сотрудников"
    ws["A15"].font = bold
    ws["A16"] = "Сотрудник"
    ws["B16"] = "Cases"
    ws["C16"] = "Ср. ответ"
    ws["D16"] = "SLA %"
    for col in ("A", "B", "C", "D"):
        ws[f"{col}16"].font = bold

    r = 17
    for emp in data["emp_kpi"]:
        ws[f"A{r}"] = emp["name"]
        ws[f"B{r}"] = emp["count"]
        ws[f"C{r}"] = emp["avg_response"]
        ws[f"D{r}"] = emp["sla"]
        r += 1

    r += 1
    ws[f"A{r}"] = "ТОП мерчантов"
    ws[f"A{r}"].font = bold
    r += 1
    ws[f"A{r}"] = "Merchant"
    ws[f"B{r}"] = "Сообщений"
    ws[f"C{r}"] = "Cases"
    ws[f"A{r}"].font = bold
    ws[f"B{r}"].font = bold
    ws[f"C{r}"].font = bold
    r += 1
    for m in data["top_merchants"]:
        ws[f"A{r}"] = m["name"]
        ws[f"B{r}"] = m["messages"]
        ws[f"C{r}"] = m["cases"]
        r += 1

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"MerchantCare_Report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/lunch/start")
async def lunch_start(request: Request):
    if _get_role(request) != "admin":
        return RedirectResponse(url="/", status_code=303)
    sheets.connect()
    now = datetime.now(TZ).strftime("%d.%m.%Y %H:%M:%S")
    sheets.set_lunch_start(now)
    return RedirectResponse(url="/", status_code=303)


@app.post("/lunch/end")
async def lunch_end(request: Request):
    if _get_role(request) != "admin":
        return RedirectResponse(url="/", status_code=303)
    sheets.connect()
    now = datetime.now(TZ).strftime("%d.%m.%Y %H:%M:%S")
    sheets.set_lunch_end(now)
    return RedirectResponse(url="/", status_code=303)


if __name__ == "__main__":
    uvicorn.run(
        "dashboard.app:app",
        host=settings.dashboard_host,
        port=settings.dashboard_port,
        reload=settings.debug,
    )